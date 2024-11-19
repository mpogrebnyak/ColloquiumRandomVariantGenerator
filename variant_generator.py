import random
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def read_students(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        students = [' '.join(line.strip().split()) for line in file.readlines() if line.strip()]
    return sorted(students)


def read_questions_from_excel(file_path):
    df = pd.read_excel(file_path, header=None)
    return {index + 1: row[0] for index, row in df.iterrows()}


def generate_student_questions(students, total_questions, questions_per_student):
    if len(students) * questions_per_student > total_questions * 3:
        raise ValueError("Недостаточно вопросов для всех студентов с учетом повторений.")

    questions = list(range(1, total_questions + 1))
    question_count = defaultdict(int)
    students_questions = defaultdict(list)

    for student in students:
        assigned_questions = set()
        while len(assigned_questions) < questions_per_student:
            question = random.choice(questions)
            if question_count[question] < 3 and question not in assigned_questions:
                assigned_questions.add(question)
                question_count[question] += 1
        students_questions[student] = list(assigned_questions)

    return students_questions


def save_to_excel(students_questions, file_path):
    df = pd.DataFrame.from_dict(students_questions, orient='index',
                                columns=[f'Question {i + 1}' for i in range(len(next(iter(students_questions.values()))))])
    df.index.name = 'Student'
    df.reset_index(inplace=True)

    df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active

    # Formatting
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill

    wb.save(file_path)
